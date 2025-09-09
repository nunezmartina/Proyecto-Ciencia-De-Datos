from airflow import DAG
from airflow.operators.python import PythonOperator
from datetime import datetime, timedelta
import pandas as pd
import os
import io
import zipfile
import requests
from typing import List

# ===================== Configuración =====================
DATA_DIR = "/tmp/movies"
RAW_PATH = f"{DATA_DIR}/raw"
STAGING_PATH = f"{DATA_DIR}/staging"
OUTPUT_PATH = "/usr/local/airflow/include/output"
MOVIELENS_URL = "http://files.grouplens.org/datasets/movielens/ml-latest-small.zip"

EXCEL_SHEET_CLEAN = "Películas (clean)"
EXCEL_SHEET_FINAL = "Películas"
EXCEL_FILENAME_FINAL = "movies_with_ratings.xlsx"
MAX_GENRE_COLS = 6  # columnas genero_1..genero_6

default_args = {
    "owner": "grupo",
    "start_date": datetime(2025, 9, 1),
    "retries": 1,
    "retry_delay": timedelta(minutes=2),
    "depends_on_past": False,
}

# ===================== Utilidades =====================
def _ensure_dirs():
    os.makedirs(RAW_PATH, exist_ok=True)
    os.makedirs(STAGING_PATH, exist_ok=True)
    os.makedirs(OUTPUT_PATH, exist_ok=True)

def _split_genres_to_cols(genres: str, max_cols: int) -> List[str]:
    if not isinstance(genres, str) or not genres.strip():
        return [""] * max_cols
    parts = [p.strip() for p in genres.split("|") if p.strip()]
    seen = set()
    uniq = [p for p in parts if not (p in seen or seen.add(p))]
    return (uniq + [""] * max_cols)[:max_cols]

def _write_excel_table_openpyxl(df: pd.DataFrame, path: str, sheet_name: str):
    """
    Escribe un .xlsx bonito con openpyxl:
    - Tabla con filtros
    - Encabezado congelado
    - Anchos de columna
    """
    from openpyxl import Workbook
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.worksheet.table import Table, TableStyleInfo
    from openpyxl.utils import get_column_letter

    if os.path.exists(path):
        os.remove(path)

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    # volcar DataFrame
    for row in dataframe_to_rows(df, index=False, header=True):
        ws.append(row)

    last_row = ws.max_row
    last_col = ws.max_column
    ref = f"A1:{get_column_letter(last_col)}{last_row}"

    tab = Table(displayName="TablaPeliculas", ref=ref)
    tab.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium9", showRowStripes=True, showColumnStripes=False
    )
    ws.add_table(tab)

    ws.freeze_panes = "A2"

    for col_idx in range(1, last_col + 1):
        col_letter = get_column_letter(col_idx)
        max_len = max((len(str(c.value)) if c.value is not None else 0) for c in ws[col_letter])
        ws.column_dimensions[col_letter].width = min(max(10, max_len + 2), 60)

    wb.save(path)

def _safe_write_excel(df: pd.DataFrame, path: str, sheet_name: str):
    """
    Intenta openpyxl con formato de tabla; si falla, usa xlsxwriter; si no hay engine, CSV fallback.
    """
    # 1) openpyxl con formato lindo
    try:
        _write_excel_table_openpyxl(df, path, sheet_name)
        return {"format": "xlsx-openpyxl", "path": path}
    except Exception:
        pass

    # 2) pandas + xlsxwriter con auto-anchos
    try:
        with pd.ExcelWriter(path, engine="xlsxwriter") as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            worksheet = writer.sheets[sheet_name]
            for i, col in enumerate(df.columns):
                max_len = max([len(str(col))] + [len(str(v)) for v in df[col].astype(str).values])
                worksheet.set_column(i, i, min(max(10, max_len + 2), 60))
        return {"format": "xlsx-xlsxwriter", "path": path}
    except Exception:
        pass

    # 3) fallback CSV con ';' (Excel AR/ES lo abre en columnas)
    csv_path = os.path.splitext(path)[0] + ".csv"
    df.to_csv(csv_path, index=False, sep=";")
    return {"format": "csv-fallback", "path": csv_path}

# ===================== Tareas =====================
def extract_data():
    _ensure_dirs()
    resp = requests.get(MOVIELENS_URL, timeout=60)
    if resp.status_code != 200:
        raise Exception(f"Error al descargar datos: {resp.status_code}")
    with zipfile.ZipFile(io.BytesIO(resp.content)) as z:
        z.extractall(RAW_PATH)

def load_raw():
    _ensure_dirs()
    src = os.path.join(RAW_PATH, "ml-latest-small", "movies.csv")
    dst = os.path.join(STAGING_PATH, "movies_raw.csv")
    pd.read_csv(src).to_csv(dst, index=False)

def transform_data():
    _ensure_dirs()
    staging_file = os.path.join(STAGING_PATH, "movies_raw.csv")
    df = pd.read_csv(staging_file)

    # Año y título limpio
    df["year"] = pd.to_numeric(df["title"].str.extract(r"\((\d{4})\)")[0], errors="coerce").astype("Int64")
    df["title_clean"] = (
        df["title"].str.replace(r"\(\d{4}\)", "", regex=True)
                   .str.replace(r"\s+", " ", regex=True)
                   .str.strip()
    )

    # Normalizar géneros
    def clean_genres(g):
        if pd.isna(g): return ""
        s = str(g).strip()
        if s.lower() == "(no genres listed)": return ""
        parts = [p.strip() for p in s.split("|") if p.strip()]
        seen = set()
        uniq = [p for p in parts if not (p in seen or seen.add(p))]
        return "|".join(uniq)

    df["genres"] = df["genres"].map(clean_genres)

    # Expandir géneros a columnas genero_1..genero_n
    genre_cols = [f"genero_{i}" for i in range(1, MAX_GENRE_COLS + 1)]
    genre_df = pd.DataFrame(
        df["genres"].apply(lambda s: _split_genres_to_cols(s, MAX_GENRE_COLS)).tolist(),
        columns=genre_cols
    )

    # Ensamble limpio (sin listas en celdas)
    df_clean = (
        pd.concat([df[["movieId", "title_clean", "year", "genres"]], genre_df], axis=1)
          .sort_values(["title_clean","year","movieId"])
          .drop_duplicates(subset=["movieId"])
          .reset_index(drop=True)
    )

    df_clean.to_csv(os.path.join(STAGING_PATH, "movies_clean.csv"), index=False)

def transform_ratings():
    _ensure_dirs()
    src = os.path.join(RAW_PATH, "ml-latest-small", "ratings.csv")
    df = pd.read_csv(src)

    stats = df.groupby("movieId").agg(
        n_ratings=("rating", "count"),
        mean_rating=("rating", "mean")
    ).reset_index()

    # Popularidad bayesiana (suavizado)
    global_mean = df["rating"].mean()
    m = 50
    stats["popularity_score"] = (
        (stats["n_ratings"] / (stats["n_ratings"] + m)) * stats["mean_rating"]
        + (m / (stats["n_ratings"] + m)) * global_mean
    )

    stats.to_csv(os.path.join(STAGING_PATH, "ratings_stats.csv"), index=False)

def join_datasets():
    _ensure_dirs()
    movies_file = os.path.join(STAGING_PATH, "movies_clean.csv")
    ratings_file = os.path.join(STAGING_PATH, "ratings_stats.csv")

    movies = pd.read_csv(movies_file)
    ratings = pd.read_csv(ratings_file)

    merged = movies.merge(ratings, on="movieId", how="left")
    for c in ["n_ratings","mean_rating","popularity_score"]:
        if c in merged.columns:
            merged[c] = merged[c].fillna(0)

    merged.to_csv(os.path.join(STAGING_PATH, "movies_with_ratings.csv"), index=False)

def export_movies_clean():
    """
    Genera OUTPUT_PATH/movies_clean.xlsx con una fila por película y columnas reales.
    Si no hay engine de Excel, deja CSV con ';' (no rompe el DAG).
    """
    _ensure_dirs()
    src = os.path.join(STAGING_PATH, "movies_clean.csv")
    if not os.path.exists(src):
        raise FileNotFoundError("No existe movies_clean.csv. Corré transform_data primero.")

    df = pd.read_csv(src)

    ordered_cols = [
        "movieId", "title_clean", "year",
        "genero_1","genero_2","genero_3","genero_4","genero_5","genero_6",
        "genres"
    ]
    ordered_cols = [c for c in ordered_cols if c in df.columns]
    df = df[ordered_cols].rename(columns={
        "movieId": "movie_id",
        "title_clean": "titulo",
        "year": "anio",
        "genres": "generos_todos",
    }).sort_values(["titulo","anio","movie_id"]).reset_index(drop=True)

    out_path = os.path.join(OUTPUT_PATH, "movies_clean.xlsx")
    _ = _safe_write_excel(df, out_path, EXCEL_SHEET_CLEAN)

def export_data():
    """
    Exporta a EXCEL (.xlsx) si hay engine disponible (openpyxl/xlsxwriter).
    Si no, genera CSV con separador ';' para que Excel lo abra en columnas.
    Nunca revienta: siempre deja algún archivo en OUTPUT_PATH.
    """
    _ensure_dirs()
    src = os.path.join(STAGING_PATH, "movies_with_ratings.csv")
    df = pd.read_csv(src)

    # Renombrar/ordenar columnas
    ordered_cols = [
        "movieId", "title_clean", "year",
        "genero_1","genero_2","genero_3","genero_4","genero_5","genero_6",
        "genres", "n_ratings", "mean_rating", "popularity_score"
    ]
    ordered_cols = [c for c in ordered_cols if c in df.columns]

    df = df[ordered_cols].rename(columns={
        "movieId": "movie_id",
        "title_clean": "titulo",
        "year": "anio",
        "genres": "generos_todos",
        "n_ratings": "cantidad_ratings",
        "mean_rating": "promedio_rating",
        "popularity_score": "popularidad_bayes",
    })

    if "promedio_rating" in df.columns:
        df["promedio_rating"] = df["promedio_rating"].astype(float).round(3)
    if "popularidad_bayes" in df.columns:
        df["popularidad_bayes"] = df["popularidad_bayes"].astype(float).round(3)

    df = df.sort_values(["titulo","anio","movie_id"]).reset_index(drop=True)

    excel_path = os.path.join(OUTPUT_PATH, EXCEL_FILENAME_FINAL)
    _ = _safe_write_excel(df, excel_path, EXCEL_SHEET_FINAL)

# ===================== DAG =====================
with DAG(
    dag_id="movies_pipeline_excel_ok",
    default_args=default_args,
    description="Películas limpias y con ratings a Excel (una fila por película, columnas limpias)",
    schedule=None,  # en Airflow<2.8 usar 'schedule_interval=None'
    catchup=False,
    tags=["movies","excel","etl"],
) as dag:

    t1 = PythonOperator(task_id="extract_data", python_callable=extract_data)
    t2 = PythonOperator(task_id="load_raw", python_callable=load_raw)
    t3 = PythonOperator(task_id="transform_data", python_callable=transform_data)
    t4 = PythonOperator(task_id="transform_ratings", python_callable=transform_ratings)
    t5 = PythonOperator(task_id="join_datasets", python_callable=join_datasets)

    # Exports
    t6_clean = PythonOperator(task_id="export_movies_clean", python_callable=export_movies_clean)
    t6_final = PythonOperator(task_id="export_data", python_callable=export_data)

    # Dependencias
    t1 >> t2
    t2 >> t3
    t1 >> t4
    t3 >> t6_clean
    [t3, t4] >> t5 >> t6_final
